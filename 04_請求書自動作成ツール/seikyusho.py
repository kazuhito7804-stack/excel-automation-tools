"""請求書自動作成ツール

会社名・品目・単価・数量が並んだ一覧表Excel(INPUT_FILE)を読み込み、
会社ごとに1シートの請求書を作成して1つのExcelファイル(OUTPUT_FILE)に出力する。

各シートの構成:
    1行目: タイトル(「請求書」)
    3行目: 宛名(会社名+御中)
    4行目: 請求日
    6行目: 見出し(品目・単価・数量・金額)
    7行目〜: 明細(品目ごとに1行、金額は単価×数量の数式)
    明細の1行空けたあと: 小計・消費税(10%)・合計金額(いずれもExcel数式)

実行方法:
    python seikyusho.py
"""

import unicodedata
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from collections import OrderedDict
from datetime import date

INPUT_FILE = "請求データ/seikyusaki.xlsx"  # 入力: 会社名・品目・単価・数量の一覧表
OUTPUT_FILE = "seikyusho_kekka.xlsx"  # 出力: 会社ごとにシート分けした請求書

# ---- 見た目に関する設定値 ----------------------------------------------

TITLE_FONT = Font(name="游ゴシック", size=16, bold=True)                 # 「請求書」タイトル用
HEADER_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")  # 見出し行(白文字)
NORMAL_FONT = Font(name="游ゴシック", size=11)                            # 通常セル
TOTAL_FONT = Font(name="游ゴシック", size=11, bold=True)                  # 合計金額行(太字)

CURRENCY_FMT = "¥#,##0"  # 金額の表示形式(桁区切り+円マーク)

HEADER_FILL = PatternFill(fill_type="solid", start_color="4472C4", end_color="4472C4")  # 見出し行の背景(青)
TOTAL_FILL = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")   # 合計金額行の背景(薄灰)

THIN = Side(style="thin", color="808080")     # 表内の細い罫線
MEDIUM = Side(style="medium", color="000000")  # 合計金額行を強調する太めの罫線
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)  # 表の各セルに使う四方罫線

RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")   # 金額・数値用(右寄せ)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")  # 見出し・数量用(中央寄せ)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")     # 品目名用(左寄せ)

# 列幅の下限(文字数ベース)。実際の内容がこれより長い場合はそちらを優先する
MIN_WIDTHS = {"A": 14, "B": 10, "C": 8, "D": 12}


def visual_width(text):
    """文字列の表示幅を概算する。

    全角文字(漢字・ひらがな・カタカナなど)は半角文字の2倍の幅として扱う。
    Excelの列幅は半角文字基準のため、そのままlen()を使うと全角文字の多い
    文字列(会社名など)で列が狭くなりすぎてしまう。そのための簡易補正。
    """
    width = 0
    for ch in str(text):
        width += 2 if unicodedata.east_asian_width(ch) in ("W", "F", "A") else 1
    return width


def load_items(path):
    """入力Excel(会社名・品目・単価・数量)を読み込み、会社名ごとにまとめる。

    Args:
        path: 入力ファイルのパス。1行目はヘッダー行(会社名・品目・単価・数量)。

    Returns:
        OrderedDict[str, list[tuple]]:
            キーが会社名、値が (品目, 単価, 数量) のタプルのリスト。
            入力ファイルに登場した順番を保つ(OrderedDictのため)。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    companies = OrderedDict()
    # 2行目以降(ヘッダーを除く)を1行ずつ読み、A〜D列(会社名・品目・単価・数量)を取得
    for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
        company, item, unit_price, qty = row
        if not company or not item:
            continue  # 空行はスキップ
        companies.setdefault(company, []).append((item, unit_price, qty))
    return companies


def build_invoice_sheet(ws, company, items):
    """1社分の請求書シートを組み立てる。

    タイトル・宛名・請求日・明細・小計/消費税/合計金額を書き込み、
    罫線・配色・列幅・行の高さ・A4印刷設定まで一括で行う。

    Args:
        ws: 書き込み対象のワークシート(空の状態で渡される)。
        company: 会社名(宛名・シート名の元になる)。
        items: (品目, 単価, 数量) のタプルのリスト。
    """
    # 各列に実際入る文字列の最大幅を記録しておき、あとで列幅に反映する
    col_max_width = dict(MIN_WIDTHS)

    def track(col_letter, text):
        """指定した列に書き込む文字列の表示幅を記録し、最大値を更新する。"""
        col_max_width[col_letter] = max(col_max_width[col_letter], visual_width(text))

    # ---- タイトル ----
    ws["A1"] = "請求書"
    ws["A1"].font = TITLE_FONT

    # ---- 宛名(会社名+御中) ----
    atena = f"{company}御中"
    ws["A3"] = atena
    ws["A3"].font = NORMAL_FONT
    track("A", atena)

    # ---- 請求日(スクリプト実行日を使用) ----
    seikyubi = f"請求日：{date.today().strftime('%Y年%m月%d日')}"
    ws["A4"] = seikyubi
    ws["A4"].font = NORMAL_FONT
    track("A", seikyubi)

    # ---- 見出し行(品目・単価・数量・金額) ----
    headers = ["品目", "単価", "数量", "金額"]
    header_row = 6
    for col_idx, text in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = GRID_BORDER
        track(col_letter, text)

    # ---- 明細行(品目ごとに1行、金額は数式で計算) ----
    first_item_row = header_row + 1
    for offset, (item, unit_price, qty) in enumerate(items):
        r = first_item_row + offset

        c_item = ws.cell(row=r, column=1, value=item)
        c_item.font = NORMAL_FONT
        c_item.alignment = LEFT_ALIGN
        c_item.border = GRID_BORDER
        track("A", item)

        c_price = ws.cell(row=r, column=2, value=unit_price)
        c_price.font = NORMAL_FONT
        c_price.number_format = CURRENCY_FMT
        c_price.alignment = RIGHT_ALIGN
        c_price.border = GRID_BORDER
        track("B", f"¥{unit_price:,}")

        c_qty = ws.cell(row=r, column=3, value=qty)
        c_qty.font = NORMAL_FONT
        c_qty.alignment = CENTER_ALIGN
        c_qty.border = GRID_BORDER
        track("C", qty)

        # 金額はハードコードせず「=単価セル*数量セル」の数式にする。
        # こうすることで、あとから単価や数量を書き換えても自動で再計算される。
        c_amount = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
        c_amount.font = NORMAL_FONT
        c_amount.number_format = CURRENCY_FMT
        c_amount.alignment = RIGHT_ALIGN
        c_amount.border = GRID_BORDER
        track("D", f"¥{unit_price * qty:,}")

    last_item_row = first_item_row + len(items) - 1

    # ---- 小計・消費税・合計金額 ----
    # 明細の最終行から1行空けて配置する(見た目の区切りのため)
    subtotal_row = last_item_row + 2
    tax_row = subtotal_row + 1
    total_row = tax_row + 1

    def set_summary_row(row, label, formula, font, fill=None):
        """小計/消費税/合計金額の1行(C列にラベル、D列に数式)を書き込む。"""
        c_label = ws.cell(row=row, column=3, value=label)
        c_label.font = font
        c_label.alignment = RIGHT_ALIGN
        c_label.border = GRID_BORDER
        if fill:
            c_label.fill = fill
        track("C", label)

        c_value = ws.cell(row=row, column=4, value=formula)
        c_value.font = font
        c_value.number_format = CURRENCY_FMT
        c_value.alignment = RIGHT_ALIGN
        c_value.border = GRID_BORDER
        if fill:
            c_value.fill = fill

    # 小計 = 明細の金額列(D列)の合計
    set_summary_row(subtotal_row, "小計", f"=SUM(D{first_item_row}:D{last_item_row})", NORMAL_FONT)
    # 消費税(10%) = 小計 × 0.1
    set_summary_row(tax_row, "消費税(10%)", f"=D{subtotal_row}*0.1", NORMAL_FONT)
    # 合計金額 = 小計 + 消費税
    set_summary_row(total_row, "合計金額", f"=D{subtotal_row}+D{tax_row}", TOTAL_FONT, fill=TOTAL_FILL)

    # 合計金額行だけ上下を太めの罫線にして目立たせる
    for col in range(1, 5):
        cell = ws.cell(row=total_row, column=col)
        border = cell.border
        cell.border = Border(left=border.left, right=border.right, top=MEDIUM, bottom=MEDIUM)

    # ---- 列幅(内容の実測幅+余白と、あらかじめ決めた最低幅の大きい方を採用) ----
    for col, width in col_max_width.items():
        ws.column_dimensions[col].width = width + 3

    # ---- 行の高さ(印刷時に詰まって見えないよう調整) ----
    ws.row_dimensions[1].height = 28  # タイトル
    ws.row_dimensions[3].height = 20  # 宛名
    ws.row_dimensions[4].height = 18  # 請求日
    ws.row_dimensions[header_row].height = 20
    for r in range(first_item_row, last_item_row + 1):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[subtotal_row].height = 18
    ws.row_dimensions[tax_row].height = 18
    ws.row_dimensions[total_row].height = 20

    # ---- A4用紙1ページに収まるよう印刷設定 ----
    ws.sheet_view.showGridLines = False  # 印刷時に不要なグリッド線を消す(罫線のみ表示)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # Excel側にも「1ページに収める」設定を反映
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_options.horizontalCentered = True  # 印刷時に表を用紙の水平方向中央に配置
    ws.page_margins = PageMargins(
        left=0.7, right=0.7, top=0.9, bottom=0.75, header=0.3, footer=0.3
    )
    ws.print_area = f"A1:D{total_row}"  # 印刷範囲をタイトル〜合計金額行までに限定


def main():
    """入力ファイルを読み込み、会社ごとの請求書シートを作って出力ファイルに保存する。"""
    companies = load_items(INPUT_FILE)

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)  # openpyxlが自動生成する空シートを削除しておく

    for company, items in companies.items():
        # Excelのシート名は31文字までという制限があるため、超える場合は切り詰める
        safe_name = company[:31]
        ws = out_wb.create_sheet(title=safe_name)
        build_invoice_sheet(ws, company, items)

    out_wb.save(OUTPUT_FILE)
    print("請求書を作成しました")


if __name__ == "__main__":
    main()
