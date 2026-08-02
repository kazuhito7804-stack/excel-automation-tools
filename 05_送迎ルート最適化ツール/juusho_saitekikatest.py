"""
児童デイサービス送迎ルート最適化プログラム
------------------------------------------
住所録(Excel)を読み込み、施設からの直線距離が近い順に
訪問リストを並べ替えて、結果を同じExcelの別シートに書き出す。

使い方:
1. 下の「設定」セクションの値を必要に応じて書き換える
2. このファイルと同じフォルダに INPUT_FILE を置く
3. 実行する(ターミナルで python juusho_saitekikatest.py)
"""
import os
os.chdir(os.path.dirname(os.path.abspath(__file__)))
import re

import time
import unicodedata
import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import openpyxl
from geopy.distance import geodesic
from geopy.exc import GeocoderTimedOut, GeocoderUnavailable
from geopy.geocoders import Nominatim

# ==================== 設定 ====================
FACILITY_ADDRESS = "〇〇市〇〇3丁目〇〇"  # 送迎の出発地(施設の住所) ※実際に使う際は各自の施設住所に書き換えてください
INPUT_FILE = "住所録.xlsx"  # 住所録のExcelファイル
OUTPUT_SHEET_NAME = "送迎ルート結果"  # 結果を書き込むシート名
INPUT_SHEET_NAME = "Sheet1"  # 住所録が入っているシート名
COUNTRY_CODE = "jp"  # 検索を日本国内に限定(海外の同名地への誤マッチを防ぐ)
PREFECTURE_HINT = "兵庫県"  # 住所に都道府県が省略されている場合の補完に使う
MAX_REASONABLE_KM = 100  # 施設からこれ以上離れていたら誤ジオコーディングを疑う
# ================================================

geolocator = Nominatim(user_agent="juusyo_saitekika_app")

def get_display_width(text):
    """全角文字は2、半角文字は1としてカウントし、実際の見た目の幅に近づける"""
    width = 0
    for ch in str(text):
        if unicodedata.east_asian_width(ch) in ("F", "W", "A"):
            width += 2
        else:
            width += 1
    return width

def simplify_address(juusho):
    """
    住所から「最初の数字が出てくる手前まで」を切り出す。

    「丁目」という文字列を探す方式だと、丁目表記のない住所
    (例:「宝塚市園部２－８」)に対応できなかったため、
    正規表現で「先頭から、最初の数字(半角・全角どちらも)が
    現れる直前まで」を取り出す方式に変更している。

    例:
        "大阪市中央区曽根崎2丁目４－３" -> "大阪市中央区曽根崎"
        "宝塚市園部２－８"             -> "宝塚市園部"
    """
    if juusho is None:
        return None

    match = re.match(r"^[^0-9０-９]+", juusho)
    if match:
        return match.group(0)
    # 数字が全く含まれない場合はそのまま返す
    return juusho


def geocode_safe(address, retries=3):
    """
    Nominatimへの問い合わせをリトライ付きで行う。

    - 無料サービスへの連続アクセスを避けるため、毎回 time.sleep(1) を入れる
    - タイムアウトなどの一時的なエラーはリトライする
    - country_codesで検索範囲を日本国内に限定し、海外の同名地への
      誤マッチ(例:遠く離れた場所がヒットして距離が異常値になる)を防ぐ
    """
    for attempt in range(retries):
        try:
            location = geolocator.geocode(
                address, timeout=10, country_codes=COUNTRY_CODE
            )
            time.sleep(1)
            return location
        except (GeocoderTimedOut, GeocoderUnavailable):
            print(f"  ...タイムアウトのため再試行します({attempt + 1}/{retries}): {address}")
            time.sleep(2)
    return None


def geocode_with_fallback(cho_made, facility_coords):
    """
    住所をジオコーディングし、(緯度経度, 距離km, 備考) を返す。

    - 通常検索で見つからない、または距離が MAX_REASONABLE_KM を超えて
      不自然に遠い場合は、PREFECTURE_HINT を先頭に補って再検索する
    - それでも解決しない場合は None を返し、呼び出し側で「要確認」として扱う
    """
    location = geocode_safe(cho_made)

    if location is not None:
        kyori = round(
            geodesic(facility_coords, (location.latitude, location.longitude)).km, 2
        )
        if kyori <= MAX_REASONABLE_KM:
            return kyori, "OK"
        note_first = f"最初の検索結果が{kyori}kmと不自然に遠いため再検索"
    else:
        note_first = "最初の検索で見つからず再検索"

    # 都道府県を補って再検索
    retry_address = PREFECTURE_HINT + cho_made
    print(f"  ...{note_first}: {retry_address}")
    location2 = geocode_safe(retry_address)
    if location2 is not None:
        kyori2 = round(
            geodesic(facility_coords, (location2.latitude, location2.longitude)).km, 2
        )
        if kyori2 <= MAX_REASONABLE_KM:
            return kyori2, "OK(都道府県を補完)"
        return None, f"再検索後も{kyori2}kmと不自然。要手動確認"

    return None, "住所が見つかりませんでした。要手動確認"


def main():
    facility_simplified = simplify_address(FACILITY_ADDRESS)
    print(f"施設の住所: {FACILITY_ADDRESS} → {facility_simplified}")
    facility_location = geocode_safe(facility_simplified)
    if facility_location is None:
        # 施設住所も念のため都道府県を補って再試行
        facility_location = geocode_safe(PREFECTURE_HINT + facility_simplified)
    if facility_location is None:
        print("施設の住所が見つかりませんでした。FACILITY_ADDRESSを確認してください。")
        return

    facility_coords = (facility_location.latitude, facility_location.longitude)
    print(f"施設の座標: {facility_coords}")

    wb = openpyxl.load_workbook(INPUT_FILE)
    input_sheet = wb[INPUT_SHEET_NAME]

    ok_results = []  # (kyori, namae, juusho, note)
    flagged_results = []  # (namae, juusho, note)

    for row in input_sheet.iter_rows(min_row=2, values_only=True):
        namae, juusho = row[0], row[1]
        if not namae or not juusho:
            continue

        cho_made = simplify_address(juusho)
        print(f"検索中: {namae}({juusho} → {cho_made})")

        kyori, note = geocode_with_fallback(cho_made, facility_coords)

        if kyori is None:
            print(f"  → 要確認: {note}")
            flagged_results.append((namae, juusho, note))
        else:
            print(f"  → {kyori}km ({note})")
            ok_results.append((kyori, namae, juusho, note))

    ok_results.sort(key=lambda x: x[0])

  # 出力シートを準備(既存なら中身をクリアして作り直す)
    if OUTPUT_SHEET_NAME in wb.sheetnames:
        del wb[OUTPUT_SHEET_NAME]
    output_sheet = wb.create_sheet(OUTPUT_SHEET_NAME)

    headers = ["訪問順", "名前", "住所", "距離(km)", "備考"]
    today_str = datetime.date.today().strftime("%Y年%m月%d日")

    # --- タイトル行(1行目) ---
    output_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = output_sheet.cell(row=1, column=1, value=f"送迎ルート結果(作成日:{today_str})")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    output_sheet.row_dimensions[1].height = 28

    # --- 見出し行(2行目) ---
    header_row_num = 2
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = output_sheet.cell(row=header_row_num, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    output_sheet.row_dimensions[header_row_num].height = 22

    # --- 罫線定義 ---
    thin = Side(style="thin", color="B0B0B0")
    thick = Side(style="medium", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- データ行(3行目〜) ---
    flagged_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    current_row = header_row_num + 1

    for order, (kyori, namae, juusho, note) in enumerate(ok_results, start=1):
        values = [order, namae, juusho, kyori, note]
        for col_idx, value in enumerate(values, start=1):
            cell = output_sheet.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 4:  # 距離(km)
                cell.number_format = "0.00"
        output_sheet.row_dimensions[current_row].height = 18
        current_row += 1

    for namae, juusho, note in flagged_results:
        values = ["要確認", namae, juusho, None, note]
        for col_idx, value in enumerate(values, start=1):
            cell = output_sheet.cell(row=current_row, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = flagged_fill
            cell.alignment = Alignment(vertical="center")
        output_sheet.row_dimensions[current_row].height = 18
        current_row += 1

    last_data_row = current_row - 1

    # 見出し行の下だけ太線に
    for col_idx in range(1, len(headers) + 1):
        output_sheet.cell(row=header_row_num, column=col_idx).border = Border(
            left=thin, right=thin, top=thin, bottom=thick
        )

    # --- 列幅の自動調整 ---
    col_widths = [get_display_width(h) for h in headers]
    for row in output_sheet.iter_rows(min_row=header_row_num, max_row=last_data_row,
                                       min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.value is not None:
                length = get_display_width(cell.value)
                col_idx = cell.column - 1
                if length > col_widths[col_idx]:
                    col_widths[col_idx] = length
    for col_idx, width in enumerate(col_widths, start=1):
        output_sheet.column_dimensions[get_column_letter(col_idx)].width = width + 4

    # --- ウィンドウ枠固定(見出し行の下から) ---
    output_sheet.freeze_panes = f"A{header_row_num + 1}"

    # --- 印刷設定(A4横・1ページ幅) ---
    output_sheet.page_setup.orientation = "landscape"
    output_sheet.page_setup.fitToWidth = 1
    output_sheet.page_setup.fitToHeight = 0
    output_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    output_sheet.print_area = f"A1:{get_column_letter(len(headers))}{last_data_row}"

    wb.save(INPUT_FILE)
    print(
        f"\n完了: {len(ok_results)}件を距離順に書き出しました"
        f"(要確認: {len(flagged_results)}件)。"
    )


if __name__ == "__main__":
    main()
