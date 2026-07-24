"""
フォーマット統一ツール
--------------
「顧客データ」フォルダの中にある、会社ごとのExcelファイル
(例:A社_顧客リスト.xlsx)を全部読み込んで、
列の並び順がバラバラでも、統一されたフォーマットの
1つのExcelファイルにまとめるプログラム。

前提となるファイルの形式:
  各ファイルの1行目はヘッダー行で、以下の4項目が含まれる。
  「会社名」「担当者名」「電話番号」「メールアドレス」
  ただし、会社ごとにこの4項目の並び順(列の位置)が違ってもよい。

出力される顧客リスト_統一.xlsxには、統一フォーマット
(会社名, 担当者名, 電話番号, メールアドレス の順)で
全社分のデータがまとまる。
"""

import glob
from openpyxl import Workbook, load_workbook

# 統一後の列の並び順。ここを変えれば、出力の列順を変更できる
UNIFIED_COLUMNS = ["会社名", "担当者名", "電話番号", "メールアドレス"]


def get_column_map(header_row):
    """
    ヘッダー行(セルのリスト)から、{項目名: 列番号} の辞書を作る関数。

    例: ヘッダーが [担当者名, 会社名, メールアドレス, 電話番号] の場合
        -> {"担当者名": 0, "会社名": 1, "メールアドレス": 2, "電話番号": 3}

    これにより、会社ごとに列の並び順が違っても、
    「項目名を指定するだけで正しい列番号が分かる」ようになる。
    """
    column_map = {}

    # enumerate() で、リストの中身と同時に「その位置(番号)」も取得する
    for index, cell in enumerate(header_row):
        column_map[cell.value] = index

    return column_map


def get_customer_data(ws, column_map):
    """
    シートの2行目以降を、column_mapを使って
    (会社名, 担当者名, 電話番号, メールアドレス) の順で
    取り出したタプルのリストにして返す関数。

    ヘッダーの並び順が会社ごとに違っていても、column_mapのおかげで
    常に同じ順番(統一フォーマット)で取り出せる。
    """
    result = []

    # 2行目から最終行まで、1行ずつ読み込む(1行目はヘッダーなので飛ばす)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # column_map["会社名"] が返す番号(例:0)を使って、
        # row[0] のように正しい位置から値を取り出す
        company = row[column_map["会社名"]]
        person = row[column_map["担当者名"]]
        tel = row[column_map["電話番号"]]
        mail = row[column_map["メールアドレス"]]

        result.append((company, person, tel, mail))

    return result


def load_all_customers(folder_path):
    """
    指定したフォルダの中にある全Excelファイルを読み込んで、
    統一フォーマットの顧客データを1つのリストにまとめて返す関数。
    """
    files = glob.glob(f"{folder_path}/*.xlsx")

    # 全社分の統一フォーマットデータを、まとめて貯めるリスト
    all_customers = []

    for file_path in files:
        wb = load_workbook(file_path)
        ws = wb.active

        # このファイルのヘッダー行を読んで、列の対応表を作る
        header_row = list(ws[1])
        column_map = get_column_map(header_row)

        # 対応表を使って、統一フォーマットでデータを取り出す
        data = get_customer_data(ws, column_map)

        # extend()で、このファイルのデータを全体のリストに追加する
        # (append()だと「リストのリスト」になってしまうので注意)
        all_customers.extend(data)

    return all_customers


def write_unified_sheet(ws, all_customers):
    """
    統一フォーマットの顧客データを、シートに書き込む関数。
    """
    # ヘッダー行を書き込む
    for col_num, column_name in enumerate(UNIFIED_COLUMNS, start=1):
        ws.cell(row=1, column=col_num, value=column_name)

    # データ行を書き込む
    for row_num, customer in enumerate(all_customers, start=2):
        for col_num, value in enumerate(customer, start=1):
            ws.cell(row=row_num, column=col_num, value=value)


def main():
    # ここを変えれば、別のフォルダのデータも統一できる
    folder_path = "顧客データ"
    output_filename = "顧客リスト_統一.xlsx"

    # 全ファイルを読み込んで、統一フォーマットのリストを作る
    all_customers = load_all_customers(folder_path)

    # 出力用の新しいワークブックを作成
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "顧客リスト"

    write_unified_sheet(ws_out, all_customers)

    wb_out.save(output_filename)
    print(f"{output_filename} を作成しました({len(all_customers)}件のデータ)")


# このファイルが直接実行されたときだけ main() を呼び出す
# (他のファイルからimportされたときは実行されない)
if __name__ == "__main__":
    main()
