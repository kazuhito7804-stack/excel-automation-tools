"""
勤怠集計ツール
--------------
「勤怠データ」フォルダの中にある、社員ごとのExcelファイル
(例:田中_7月勤怠.xlsx)を全部読み込んで、
社員ごとの月間残業時間を集計するプログラム。

前提となるファイルの形式:
  ファイル名: 「社員名_○月勤怠.xlsx」
  中身      : 日付・出勤時刻・退勤時刻の3列(1行目はヘッダー)

定時は 9:00〜17:00(休憩1時間を含む)。
つまり、実際の実働時間は 8時間 - 1時間 = 7時間。
これを超えた分を残業時間として扱う。

出力される集計結果.xlsxには、2つのシートができる。
  1. 集計結果     … 社員×月ごとの総残業時間一覧
  2. 社員別合計   … 社員ごとの総残業時間(月をまたいでも合算)
"""

import glob
import os
from openpyxl import Workbook, load_workbook

# 休憩時間(分)。9:00〜17:00の間に、昼休憩が1時間ある想定
BREAK_MIN = 60

# 定時の実働時間(分)。8時間(9:00〜17:00)から休憩1時間を引いた分
TEIJI_MIN = 7 * 60


def get_employee_and_month(filename):
    """
    ファイル名から「社員名」と「月」を取り出す関数。

    例: "田中_7月勤怠.xlsx" を渡すと -> ("田中", "7") を返す

    ファイル名は "社員名_○月勤怠.xlsx" という形式である前提。
    （売上集計ツールの get_store_and_month と同じ考え方）
    """
    # "_" で分割して、社員名の部分(0番目)と残りの部分(1番目)に分ける
    parts = filename.split("_")
    employee = parts[0]

    # 残りの部分("7月勤怠.xlsx")から、さらに "月" で分割して
    # 数字の部分だけ(0番目)を取り出す
    month_part = parts[1]
    month = month_part.split("月")[0]

    return employee, month


def calc_zangyo(start, end):
    """
    出勤時刻・退勤時刻(文字列)から、その日の残業時間(分)を計算する関数。

    例: calc_zangyo("9:00", "18:30") -> 90(分)

    計算の流れ:
      1. "9:00" のような文字列を、時と分に分解する
      2. 分解した値を int() で数字に変換する
      3. 時刻をすべて「分」に統一する(時×60+分)
      4. 退勤の分 - 出勤の分 = 実働時間(休憩を含む)
      5. そこから休憩時間(60分)を引く
      6. 定時分(420分)を超えた分を残業として返す
         (マイナスになる場合=早退等は、残業0として扱う)
    """
    # ":" で分割して、時と分をそれぞれ文字列として取り出す
    start_hour, start_min = start.split(":")
    end_hour, end_min = end.split(":")

    # 文字列のままだと計算できないので、int() で数字に変換する
    start_hour = int(start_hour)
    start_min = int(start_min)
    end_hour = int(end_hour)
    end_min = int(end_min)

    # 出勤・退勤時刻を、それぞれ「分」に統一する
    # (例: 9時0分 → 9×60+0 = 540分)
    start_total_min = start_hour * 60 + start_min
    end_total_min = end_hour * 60 + end_min

    # 退勤 - 出勤 = 実働時間(休憩を含む、まだ引いていない状態)
    worked_min = end_total_min - start_total_min

    # 休憩時間を引いて、本当の実働時間を出す
    actual_worked_min = worked_min - BREAK_MIN

    # 定時分(7時間=420分)を超えた分が残業時間
    zangyo_min = actual_worked_min - TEIJI_MIN

    # 早退などでマイナスになった場合は、残業なし(0分)として扱う
    # (早退した分を、他の日の残業と相殺しないようにするため)
    if zangyo_min < 0:
        zangyo_min = 0

    return zangyo_min


def load_all_attendance(folder_path):
    """
    指定したフォルダの中にある全Excelファイルを読み込んで、
    (社員名, 月) をキーにした「総残業時間(分)」の辞書を作って返す関数。
    """
    # フォルダ内の.xlsxファイル一覧を取得
    # (返ってくるのは "勤怠データ/田中_7月勤怠.xlsx" のような、
    #  フォルダ名込みのパス)
    files = glob.glob(os.path.join(folder_path, "*.xlsx"))

    # 集計結果を貯めておく辞書。キーは (社員名, 月) のタプル
    shukei = {}

    for file_path in files:
        # フォルダ名を除いた「ファイル名だけ」を取り出す
        filename_only = os.path.basename(file_path)

        # ファイル名から社員名と月を取得
        employee, month = get_employee_and_month(filename_only)

        # Excelファイルを開いて、アクティブなシートを取得
        wb = load_workbook(file_path)
        ws = wb.active

        # その社員・その月の総残業時間を貯める変数
        # (社員が変わるたびに、ここで0にリセットされる)
        total_zangyo_min = 0

        # 2行目から最終行まで、1行ずつ読み込む(1行目はヘッダーなので飛ばす)
        # 各行は (日付, 出勤時刻, 退勤時刻) の並び
        for row in ws.iter_rows(min_row=2, values_only=True):
            date, start, end = row

            # その日の残業時間を計算して、月の合計に足し込む
            zangyo = calc_zangyo(start, end)
            total_zangyo_min += zangyo

        # 社員・月の組み合わせをキーにして、総残業時間を記録
        key = (employee, month)
        shukei[key] = total_zangyo_min

    return shukei


def write_detail_sheet(ws, shukei):
    """
    「集計結果」シートに、社員×月ごとの総残業時間一覧を書き込む関数。
    """
    ws["A1"] = "社員名"
    ws["B1"] = "月"
    ws["C1"] = "総残業時間(分)"

    row_num = 2
    for key, value in shukei.items():
        employee, month = key
        ws[f"A{row_num}"] = employee
        ws[f"B{row_num}"] = month
        ws[f"C{row_num}"] = value
        row_num += 1


def write_employee_summary_sheet(ws, shukei):
    """
    「社員別合計」シートに、社員ごとの総残業時間(月をまたいでも合算)を書き込む関数。
    """
    employee_totals = {}

    for (employee, month), zangyo in shukei.items():
        # .get(employee, 0) は「まだキーがなければ0を返す」という意味。
        # 初めて出てきた社員でもエラーにならずに加算できる。
        employee_totals[employee] = employee_totals.get(employee, 0) + zangyo

    ws["A1"] = "社員名"
    ws["B1"] = "総残業時間(分)"

    row_num = 2
    for employee, total in employee_totals.items():
        ws[f"A{row_num}"] = employee
        ws[f"B{row_num}"] = total
        row_num += 1


def main():
    # ここを変えれば、別のフォルダのデータも集計できる
    folder_path = "勤怠データ"
    output_filename = "勤怠集計結果.xlsx"

    # 全ファイルを読み込んで集計する
    shukei = load_all_attendance(folder_path)

    # 出力用の新しいワークブックを作成
    wb_out = Workbook()

    # 1つ目のシート(最初から存在するシート)を「集計結果」として使う
    ws_detail = wb_out.active
    ws_detail.title = "集計結果"
    write_detail_sheet(ws_detail, shukei)

    # 2つ目のシート「社員別合計」を新しく追加
    ws_employee = wb_out.create_sheet("社員別合計")
    write_employee_summary_sheet(ws_employee, shukei)

    wb_out.save(output_filename)
    print(f"{output_filename} を作成しました(シート:集計結果・社員別合計)")


# このファイルが直接実行されたときだけ main() を呼び出す
# (他のファイルからimportされたときは実行されない)
if __name__ == "__main__":
    main()
