"""
売上集計ツール
--------------
「売上データ」フォルダの中にある、店舗ごと・月ごとのExcelファイル
(例:大阪支店_9月売上実績.xlsx)を全部読み込んで、
1つの集計結果Excelファイルにまとめるプログラム。

出力される集計結果.xlsxには、3つのシートができる。
  1. 集計結果   … 店舗×月×商品ごとの売上一覧
  2. 店舗別合計 … 店舗ごとの売上合計
  3. 商品別合計 … 商品ごとの売上合計
"""

import glob
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


def style_header_row(ws, num_columns):
    """
    1行目(見出し行)を太字にする関数。
    num_columns には、見出しが何列あるか(例: 4列ならA〜D)を渡す。
    """
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)


def adjust_column_widths(ws):
    """
    各列の中身の長さに合わせて、列幅を自動調整する関数。
    店舗名や商品名が長くても、セルの中で見切れないようにする。
    """
    for column_cells in ws.columns:
        # その列の中で、一番長い文字列の長さを調べる
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length

        # 少し余白を持たせて列幅を設定する
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = max_length + 4


def get_store_and_month(filename):
    """
    ファイル名から「店舗名」と「月」を取り出す関数。

    例: "大阪支店_9月売上実績.xlsx" を渡すと
        -> ("大阪支店", "9") を返す

    ファイル名は "店舗名_○月売上実績.xlsx" という形式である前提。
    形式に合わない場合は、エラーで止めるのではなく (None, None) を返す。
    呼び出し側で、この (None, None) を見て「スキップ対象」と判断する。
    """
    # "_" で分割して、店舗名の部分(0番目)と残りの部分(1番目)に分ける
    parts = filename.split("_")

    # "_" が無い(=分割しても1個のままの)ファイル名は、想定形式ではない
    if len(parts) < 2:
        return None, None

    store = parts[0]

    # 残りの部分("9月売上実績.xlsx")から、さらに "月" で分割して
    # 数字の部分だけ(0番目)を取り出す
    month_part = parts[1]

    # "月" という文字が含まれていない場合も、想定形式ではない
    if "月" not in month_part:
        return None, None

    month = month_part.split("月")[0]

    # 月の部分が数字になっていない場合(例: "九月"など)も、想定形式ではない
    if not month.isdigit():
        return None, None

    return store, month


def load_all_sales(folder_path):
    """
    指定したフォルダの中にある全Excelファイルを読み込んで、
    (店舗, 月, 商品名) をキーにした売上金額の辞書を作って返す関数。

    ファイル名が想定形式と違うファイルは、集計を止めずにスキップし、
    その一覧(skipped_files)もあわせて返す。
    """
    # globで「フォルダ内の.xlsxファイル一覧」を取得する
    # (ここで返るのは "売上データ/大阪支店_9月売上実績.xlsx" のような、
    #  フォルダ名込みのパス)
    files = glob.glob(os.path.join(folder_path, "*.xlsx"))

    # 集計結果を貯めておく辞書。キーは (店舗, 月, 商品名) のタプル
    shukei = {}

    # スキップしたファイル名を貯めておくリスト
    skipped_files = []

    for file_path in files:
        # フォルダ名を除いた「ファイル名だけ」を取り出す
        filename_only = os.path.basename(file_path)

        # ファイル名から店舗名と月を取得
        store, month = get_store_and_month(filename_only)

        # ファイル名が想定形式と違った場合(store が None)は、
        # このファイルの処理を飛ばして、次のファイルに進む
        if store is None:
            skipped_files.append(filename_only)
            continue

        # Excelファイルを開いて、アクティブなシートを取得
        # data_only=True にすることで、セルに関数(例: =B2*C2)が入っていても
        # 数式そのものではなく、計算済みの結果を読み込めるようにする
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 2行目から最終行まで、1行ずつ読み込む(1行目はヘッダーなので飛ばす)
        # 各行は (商品名, 単価, 数量, 売上金額) の並び
        for row in ws.iter_rows(min_row=2, values_only=True):
            product, price, qty, amount = row

            # 店舗・月・商品の組み合わせをキーにして売上金額を記録
            # (同じ店舗・月・商品の行が複数あっても、上書きせずに加算する)
            key = (store, month, product)
            shukei[key] = shukei.get(key, 0) + amount

    return shukei, skipped_files


def write_detail_sheet(ws, shukei):
    """
    「集計結果」シートに、店舗×月×商品ごとの売上一覧を書き込む関数。
    """
    # ヘッダー行
    ws["A1"] = "店舗"
    ws["B1"] = "月"
    ws["C1"] = "商品名"
    ws["D1"] = "売上金額"

    row_num = 2
    for key, value in shukei.items():
        store, month, product = key
        ws[f"A{row_num}"] = store
        ws[f"B{row_num}"] = month
        ws[f"C{row_num}"] = product
        ws[f"D{row_num}"] = value
        # 売上金額のセルに、3桁区切り(例: 1,200,000)の書式を設定する
        ws[f"D{row_num}"].number_format = "#,##0"
        row_num += 1

    # 見出し行(1行目)を太字にする
    style_header_row(ws, num_columns=4)
    # 列幅を、中身の長さに合わせて自動調整する
    adjust_column_widths(ws)


def write_store_summary_sheet(ws, shukei):
    """
    「店舗別合計」シートに、店舗ごとの売上合計を書き込む関数。
    """
    # 店舗ごとの合計を貯める辞書
    store_totals = {}

    for (store, month, product), amount in shukei.items():
        # .get(store, 0) は「まだキーがなければ0を返す」という意味。
        # これで、初めて出てきた店舗でもエラーにならずに加算できる。
        store_totals[store] = store_totals.get(store, 0) + amount

    ws["A1"] = "店舗"
    ws["B1"] = "売上合計"

    row_num = 2
    for store, total in store_totals.items():
        ws[f"A{row_num}"] = store
        ws[f"B{row_num}"] = total
        ws[f"B{row_num}"].number_format = "#,##0"
        row_num += 1

    # 見出し行(1行目)を太字にする
    style_header_row(ws, num_columns=2)
    # 列幅を、中身の長さに合わせて自動調整する
    adjust_column_widths(ws)


def write_product_summary_sheet(ws, shukei):
    """
    「商品別合計」シートに、商品ごとの売上合計を書き込む関数。
    """
    product_totals = {}

    for (store, month, product), amount in shukei.items():
        product_totals[product] = product_totals.get(product, 0) + amount

    ws["A1"] = "商品名"
    ws["B1"] = "売上合計"

    row_num = 2
    for product, total in product_totals.items():
        ws[f"A{row_num}"] = product
        ws[f"B{row_num}"] = total
        ws[f"B{row_num}"].number_format = "#,##0"
        row_num += 1

    # 見出し行(1行目)を太字にする
    style_header_row(ws, num_columns=2)
    # 列幅を、中身の長さに合わせて自動調整する
    adjust_column_widths(ws)


def main():
    # ここを変えれば、別のフォルダのデータも集計できる
    folder_path = "売上データ"
    output_filename = "集計結果.xlsx"

    # 全ファイルを読み込んで集計する
    shukei, skipped_files = load_all_sales(folder_path)

    # ファイル名が想定形式と違ってスキップされたファイルがあれば、
    # 集計は続けつつ、最後にまとめて分かりやすく警告する
    if skipped_files:
        print("⚠ 以下のファイルは、ファイル名が「店舗名_○月売上実績.xlsx」の形式と")
        print("  異なっていたため、集計されませんでした:")
        for name in skipped_files:
            print(f"   - {name}")
        print()

    # 出力用の新しいワークブックを作成
    wb_out = Workbook()

    # 1つ目のシート(最初から存在するシート)を「集計結果」として使う
    ws_detail = wb_out.active
    ws_detail.title = "集計結果"
    write_detail_sheet(ws_detail, shukei)

    # 2つ目のシート「店舗別合計」を新しく追加
    ws_store = wb_out.create_sheet("店舗別合計")
    write_store_summary_sheet(ws_store, shukei)

    # 3つ目のシート「商品別合計」を新しく追加
    ws_product = wb_out.create_sheet("商品別合計")
    write_product_summary_sheet(ws_product, shukei)

    wb_out.save(output_filename)
    print(f"{output_filename} を作成しました(シート:集計結果・店舗別合計・商品別合計)")


# このファイルが直接実行されたときだけ main() を呼び出す
# (他のファイルからimportされたときは実行されない)
if __name__ == "__main__":
    main()
